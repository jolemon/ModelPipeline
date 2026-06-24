from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

FONT_FAMILY = "微软雅黑"
FONT_SIZE = 11


class ExcelWriter:
    """Writes structured DataFrames to formatted Excel sheets."""

    BASE_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE)
    TITLE_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, bold=True)
    HEADER_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def write(self, output_path: str, sheets: dict) -> None:
        """Write structured data to Excel.

        Args:
            output_path: Path to output .xlsx file.
            sheets: dict mapping sheet_name → dict[str, DataFrame].
        """
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for sheet_name, sections in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            current_row = 1

            for section_name, df in sections.items():
                if df is None:
                    continue
                if isinstance(df, list):
                    for label, sub_df in df:
                        if sub_df is None or (hasattr(sub_df, 'empty') and sub_df.empty):
                            continue
                        current_row = self._write_dataframe(ws, current_row, str(label), sub_df)
                elif hasattr(df, 'empty') and df.empty:
                    continue
                else:
                    current_row = self._write_dataframe(ws, current_row, section_name, df)

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

        wb.save(output_path)

    def _write_dataframe(self, ws, start_row: int, section_name: str, df) -> int:
        """Write a single DataFrame section to the worksheet."""
        current_row = start_row

        # Section title — bold, centered
        title_cell = ws.cell(row=current_row, column=1, value=section_name)
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

        # Headers — white bold on blue, centered
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
        current_row += 1

        # Data — base font, centered
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = self.BASE_FONT
                cell.alignment = self.DATA_ALIGNMENT
            current_row += 1

        # Auto-fit column widths
        self._auto_fit_columns(ws, df, start_row + 1)

        # Apply data bars
        self._apply_data_bars(ws, df, start_row + 1)

        # Spacing
        current_row += 1

        return current_row

    def _auto_fit_columns(self, ws, df, header_row: int) -> None:
        """Set column widths based on content, accounting for CJK characters."""
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = self._str_width(str(col_name))
            for row_idx in range(header_row + 1, header_row + 1 + len(df)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, self._str_width(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    @staticmethod
    def _str_width(s: str) -> float:
        """Approximate display width: CJK chars count as 2, ASCII as 1."""
        w = 0.0
        for ch in s:
            if '一' <= ch <= '鿿' or '　' <= ch <= '〿' or '＀' <= ch <= '￯':
                w += 2.0
            else:
                w += 1.0
        return w

    def _apply_data_bars(self, ws, df, header_row: int) -> None:
        """Apply data bar conditional formatting to rate-like columns."""
        data_bar_patterns = ["woe", "bad_rate", "lift", "cum_lift"]

        for col_idx, col_name in enumerate(df.columns, 1):
            col_lower = col_name.lower().replace(" ", "_")
            if any(pat in col_lower for pat in data_bar_patterns):
                col_letter = get_column_letter(col_idx)
                data_start = header_row + 1
                data_end = header_row + len(df)
                if data_end >= data_start:
                    rule = DataBarRule(
                        start_type="min",
                        end_type="max",
                        color="5B9BD5",
                        showValue=True,
                    )
                    ws.conditional_formatting.add(
                        f"{col_letter}{data_start}:{col_letter}{data_end}", rule
                    )
