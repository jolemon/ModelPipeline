import pytest
import tempfile
from pathlib import Path
import pandas as pd
from model_report import ReportGenerator, ReportConfig


class TestIntegration:
    """End-to-end test using real test.csv data and mock scorecard."""

    def test_full_report_generation(self, sample_df, mock_scorecard):
        # Set var_names to match actual test.csv feature columns
        mock_scorecard.get_var_names.return_value = [
            "txriskscorev7", "rat1_sum_bu_1_14_pboc_aj65_cur",
            "cu_ln_acct_lurate_gt50_arto", "r24m_cfcbnkndpst_noacct_ln_tac",
            "sum_bu_1_01_pboc_aa1_m12m", "agaa_zbvg_xavm_bbvf_mf",
            "r12p24m_njslap_qtcnt",
        ]
        config = ReportConfig(
            partition_col="part_id",
            cust_col="appid",
            date_col="loan_date",
            target_col="mob6_30",
            flag_col="data_flag",
            score_col="score",
            sc_score_col="score",
        )

        gen = ReportGenerator(mock_scorecard, config)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            gen.to_excel(output_path, sample_df)

            assert Path(output_path).exists()
            assert Path(output_path).stat().st_size > 0

            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            assert config.sheet1_name in wb.sheetnames
            assert config.sheet2_name in wb.sheetnames
            assert config.sheet3_name in wb.sheetnames

        finally:
            Path(output_path).unlink()

    def test_report_with_test_csv(self, sample_df, mock_scorecard):
        """Verify test.csv can be processed without errors."""
        mock_scorecard.get_var_names.return_value = [
            "txriskscorev7", "rat1_sum_bu_1_14_pboc_aj65_cur",
            "cu_ln_acct_lurate_gt50_arto", "r24m_cfcbnkndpst_noacct_ln_tac",
            "sum_bu_1_01_pboc_aa1_m12m", "agaa_zbvg_xavm_bbvf_mf",
            "r12p24m_njslap_qtcnt",
        ]
        config = ReportConfig(
            partition_col="part_id",
            cust_col="appid",
            date_col="loan_date",
            target_col="mob6_30",
            flag_col="data_flag",
            score_col="score",
            sc_score_col="score",
        )

        gen = ReportGenerator(mock_scorecard, config)
        result = gen.generate(sample_df)

        assert len(result) == 3
        for sheet_name, sections in result.items():
            assert len(sections) > 0
            for section_name, df in sections.items():
                if isinstance(df, pd.DataFrame):
                    assert len(df) > 0
